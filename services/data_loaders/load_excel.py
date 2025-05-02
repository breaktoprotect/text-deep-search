from typing import List, Dict, Union
from pathlib import Path
from openpyxl import load_workbook


def list_sheets(file_path: Union[str, Path]) -> List[str]:
    """
    Return all sheet names in the Excel workbook.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    return wb.sheetnames


def list_columns(file_path: Union[str, Path], sheet_name: str) -> List[str]:
    """
    Return the column headers of the given sheet in the Excel workbook.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return [col for col in header_row if col is not None]


def extract_data(
    file_path: Union[str, Path],
    sheet_name: str,
    columns: List[str],
) -> List[Dict[str, Union[str, float]]]:
    """
    Extract rows from the given sheet using only the specified columns.
    Returns a list of dictionaries where each dict represents a row.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[sheet_name]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx_map = {name: idx for idx, name in enumerate(header) if name in columns}

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {
            col: row[idx]
            for col, idx in col_idx_map.items()
            if idx < len(row) and row[idx] is not None
        }
        if any(row_data.values()):
            data.append(row_data)
    return data
